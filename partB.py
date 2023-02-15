import openpyxl
import random
import matplotlib.pyplot as plt

# Open the Excel file
workbook = openpyxl.load_workbook("Supervisors.xlsx")

# Select the first sheet
sheet = workbook.worksheets[0]

# Initialize the dictionary
supervisors = {}

# Loop through each row and add it to the dictionary
for row in sheet.iter_rows(values_only=True):
    supervisors[row[0]] = row[1]

workbook = openpyxl.load_workbook("Student-choices.xlsx")
sheet = workbook.worksheets[0]
student_choices = {}

# Loop through each row and add it to the dictionary
for row in sheet.iter_rows(values_only=True):
    key = row[0]
    value = list(row[1:])
    student_choices[key] = value

supervisor_dict = {}

for supervisor in supervisors.keys():
    supervisor_dict[supervisor] = None

print(supervisors)
print(student_choices)


def initialize_population(population_size):
    population = []
    for i in range(population_size):
        individual = {}
        supervisor_list = list(supervisors.keys())
        student_list = list(student_choices.keys())

        for i, supervisor in enumerate(supervisor_list):
            individual[supervisor] = []
            while len(individual[supervisor]) < supervisors[supervisor]:
                new_student = random.choice(student_list)
                while new_student in individual[supervisor] or i+1 not in student_choices[new_student]:
                    new_student = random.choice(student_list)
                individual[supervisor].append(new_student)
                student_list.remove(new_student)
        population.append(individual)
    return population


def fitness(solution, students):
    # Calculate the fitness of the solution
    fitness = 0
    i = 0
    for supervisor, assigned_students in solution.items():
        i += 1
        for student in assigned_students:
            # Add the preference of the student for the assigned supervisor
            fitness += students[student].index(i) + 1
    return fitness / 46


def select_top_candidates(population, fitness_function, x):
    population_fitness = [(individual, fitness_function(individual, students=student_choices)) for individual in population]
    population_fitness.sort(key=lambda x: x[1], reverse=False)
    return [individual for (individual, fitness) in population_fitness[:x]]


def single_point_crossover(mapping1, mapping2):
    supervisors_list = list(mapping1.keys())
    student_list = list(student_choices.keys())
    used_students = set()
    child_mapping = {}

    # Choose a random crossover point
    crossover_point = random.randint(1, len(supervisors_list))

    # Copy the first part of mapping1 into the child mapping
    for i in range(crossover_point):
        supervisor = supervisors_list[i]
        child_mapping[supervisor] = mapping1[supervisor][:]
        used_students.update(mapping1[supervisor][:])

    # Copy the second part of mapping2 into the child mapping
    for i in range(crossover_point, len(supervisors_list)):
        supervisor = supervisors_list[i]
        child_mapping[supervisor] = mapping2[supervisor][:]
        used_students.update(mapping1[supervisor][:])

    # Make sure no student is assigned to more than one supervisor
    used_students = set()
    for supervisor in supervisors_list:
        for student in child_mapping[supervisor]:
            if student in used_students:
                # If the student is already used, remove it from this supervisor's list
                child_mapping[supervisor].remove(student)
            else:
                used_students.add(student)

    # Make sure every supervisor is at or below capacity
    for supervisor in supervisors_list:
        capacity = supervisors[supervisor]
        while len(child_mapping[supervisor]) > capacity:
            # If the supervisor is over capacity, remove a random student
            student = random.choice(child_mapping[supervisor])
            child_mapping[supervisor].remove(student)

    # Add unassigned students to preferred supervisors
    unassigned_students = set(student_list) - used_students
    # print(unassigned_students)
    for student in unassigned_students:
        preferred_supervisors = student_choices[student]
        for supervisor in preferred_supervisors:
            if len(child_mapping[f"Supervisor_{supervisor}"]) < supervisors[f"Supervisor_{supervisor}"]:
                child_mapping[f"Supervisor_{supervisor}"].append(student)
                break

    return child_mapping


def mutate(mapping, mutation_rate):
    #new_mapping = mapping.copy()

    # Iterate through each supervisor-student pair
    for supervisor, students in mapping.items():
        for i, student in enumerate(students):
            # Randomly decide whether to mutate the student
            if random.random() < mutation_rate:
                # Swap the student with another random student in the same supervisor's list
                j = random.randint(0, len(students)-1)
                mapping[supervisor][i], mapping[supervisor][j] = mapping[supervisor][j], mapping[supervisor][i]

    return mapping


population = initialize_population(population_size=1000)

elite_size = int(len(population) * 0.1)
print(elite_size)

generation_fitness = []
generations = 500

for i in range(generations):
    selection = select_top_candidates(population, fitness_function=fitness, x=elite_size)
    population = selection

    crossover_children = []
    for i in range(0, len(population), 2):
        crossover1 = single_point_crossover(mapping1=population[i], mapping2=population[i+1])
        crossover2 = single_point_crossover(mapping1=population[i+1], mapping2=population[i])
        crossover_children.append(crossover1)
        crossover_children.append(crossover2)

    population = selection + crossover_children
    for solution in population:
        solution = mutate(solution, 0.05)

    fitness_scores = []

    for solution in population:
        fitness_value = fitness(solution, student_choices)
        fitness_scores.append(fitness_value)

    generation_fitness.append(sum(fitness_scores) / (len(fitness_scores)))
    print(f"{sum(fitness_scores) / (len(fitness_scores))}, {len(fitness_scores)}")

example = population[random.randint(0, len(population)-1)]
print(example)
print(fitness(example, student_choices))

plt.xlabel("Generation number")
plt.ylabel("Generation_fitness")
plt.title("Generation fitness vs Generation number")

plt.plot(list(range(generations)), generation_fitness)
plt.show()
